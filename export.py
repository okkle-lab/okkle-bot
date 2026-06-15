"""CSV export and weekly summary.

The CSV deliberately carries source_type and confirmation_status columns so an
accountant can judge how reliable each row is (a typed mileage estimate is not
the same as a confirmed receipt).
"""
import csv
import io

from sqlalchemy.orm import Session

import tax
from models import Record, User

CSV_COLUMNS = [
    "date", "period_start", "period_end", "entry_frequency", "record_type",
    "vehicle_type", "platform_or_vendor", "amount_gbp", "miles", "category",
    "source_type", "confirmation_status", "included_in_total", "confidence",
    "original_file_reference", "notes",
]


def _exportable(db: Session, user_id: int, include_review: bool = False) -> list[Record]:
    """Records that count toward totals (confirmed/estimated). Pass include_review
    to also return review-only items — they belong in the export but NOT in totals."""
    statuses = ["confirmed", "estimated"]
    if include_review:
        statuses.append("review_required")
    return (
        db.query(Record)
        .filter(Record.user_id == user_id,
                Record.confirmation_status.in_(statuses))
        .order_by(Record.record_date.asc(), Record.id.asc())
        .all()
    )


def _miles_by_vehicle(rows: list[Record], default_type: str | None) -> dict[str, float]:
    """Sum confirmed mileage per vehicle type (older rows fall back to the user's)."""
    agg: dict[str, float] = {}
    for r in rows:
        if r.record_type != "mileage":
            continue
        vt = tax.normalise_vehicle(r.vehicle_type or default_type)
        agg[vt] = agg.get(vt, 0.0) + (r.miles or 0.0)
    return agg


def build_csv(db: Session, user_id: int) -> str:
    rows = _exportable(db, user_id, include_review=True)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(CSV_COLUMNS)
    for r in rows:
        in_total = "no" if r.confirmation_status == "review_required" else "yes"
        writer.writerow([
            r.record_date, r.period_start or "", r.period_end or "",
            r.entry_frequency or "", r.record_type, r.vehicle_type or "",
            r.platform_or_vendor,
            f"{r.amount:.2f}" if r.amount is not None else "",
            f"{r.miles:.1f}" if r.miles is not None else "",
            r.category, r.source_type, r.confirmation_status, in_total,
            f"{r.confidence:.2f}", r.original_media_url, r.notes,
        ])
    return buf.getvalue()


def weekly_summary(db: Session, user: User) -> str:
    """A short text recap to send over WhatsApp (the 'tax-saved counter' hook).

    Uses the user's vehicle type for the mileage rate and their tax-estimate
    level for the rough tax-benefit figure.
    """
    rows = _exportable(db, user.id)
    income = sum(r.amount or 0 for r in rows if r.record_type == "income")
    expenses = sum(r.amount or 0 for r in rows if r.record_type == "expense")

    by_vehicle = _miles_by_vehicle(rows, user.vehicle_type)
    total_miles = sum(by_vehicle.values())
    deduction = sum(tax.mileage_deduction(m, vt) for vt, m in by_vehicle.items())

    lines = [
        "Your records so far:",
        f"• Income logged: £{income:,.2f}",
        f"• Expenses logged (for accountant review): £{expenses:,.2f}",
        f"• Business miles: {total_miles:,.0f}  (≈ £{deduction:,.2f} mileage deduction)",
    ]
    # Break the mileage down per vehicle when more than one has been used.
    used = {vt: m for vt, m in by_vehicle.items() if m > 0}
    if len(used) > 1:
        for vt, m in sorted(used.items(), key=lambda x: -x[1]):
            lines.append(
                f"    – {tax.emoji(vt)} {tax.label(vt)}: {m:,.0f} mi "
                f"(£{tax.mileage_deduction(m, vt):,.2f})"
            )
    if user.tax_rate:
        benefit = tax.tax_benefit(deduction, user.tax_rate)
        lines.append(
            f"• Estimated tax benefit from mileage: up to ~£{benefit:,.2f} "
            f"(at {user.tax_rate * 100:.0f}% tax rate)"
        )
    lines.append("\nIndicative only, based on what you've confirmed — not tax advice.")
    return "\n".join(lines)


# --- Flow F: weekly / monthly courier summary --------------------------------

import datetime as _dt


def _period_range(monthly: bool, ref: "_dt.date | None" = None):
    d = ref or _dt.date.today()
    if monthly:
        start = d.replace(day=1)
        nxt = (start.replace(year=start.year + 1, month=1) if start.month == 12
               else start.replace(month=start.month + 1))
        return start, nxt - _dt.timedelta(days=1)
    start = d - _dt.timedelta(days=d.weekday())
    return start, start + _dt.timedelta(days=6)


def _fmt(d: "_dt.date") -> str:
    return f"{d.day} {d:%b %Y}"


def _streak(all_dates: list[str], monthly: bool) -> int:
    """Count consecutive periods (back from now) with at least one record."""
    count = 0
    ref = _dt.date.today()
    for _ in range(60):
        start, end = _period_range(monthly, ref)
        s, e = start.isoformat(), end.isoformat()
        if any(s <= d <= e for d in all_dates):
            count += 1
            ref = start - _dt.timedelta(days=1)
        else:
            break
    return count


def summary(db: Session, user: User, period: dict | None = None) -> str:
    """Flow F summary. Defaults to the user's current period; pass a `period`
    dict (from periods.resolve) to summarise any past/custom range."""
    monthly = (getattr(user, "log_frequency", "weekly") or "weekly") == "monthly"
    if period:
        start, end = period["start"], period["end"]
        title = period["title"]
        monthly = period["frequency"] in ("monthly", "annual")
    else:
        start, end = _period_range(monthly)
        title = "This month's" if monthly else "This week's"
    s_iso, e_iso = start.isoformat(), end.isoformat()
    period_line = f"Period: {_fmt(start)} – {_fmt(end)}"

    all_rows = _exportable(db, user.id)
    rows = [r for r in all_rows if s_iso <= (r.record_date or "") <= e_iso]
    income = [r for r in rows if r.record_type == "income"]
    mileage = [r for r in rows if r.record_type == "mileage"]
    expenses = [r for r in rows if r.record_type == "expense"]

    # Empty period (§18).
    if not rows:
        return (f"No records found for this period yet.\n\n{period_line}\n\n"
                "You can add mileage, earnings or expenses to start your record.")

    earnings_total = sum(r.amount or 0 for r in income)
    by_vehicle = _miles_by_vehicle(mileage, user.vehicle_type)
    miles_total = sum(by_vehicle.values())
    deduction = sum(tax.mileage_deduction(m, vt) for vt, m in by_vehicle.items())
    benefit = tax.tax_benefit(deduction, user.tax_rate) if user.tax_rate else 0.0
    streak = _streak([r.record_date for r in all_rows], monthly)
    streak_unit = "month" if monthly else "week"

    def _by_platform():
        agg: dict[str, float] = {}
        for r in income:
            agg[r.platform_or_vendor or "Other"] = agg.get(r.platform_or_vendor or "Other", 0.0) + (r.amount or 0)
        return sorted(agg.items(), key=lambda x: -x[1])

    # Earnings-only (§5/§6).
    if income and not mileage:
        lines = [f"{title} earnings record\n", period_line, ""]
        ranked = _by_platform()
        if len(ranked) > 1:
            for i, (name, amt) in enumerate(ranked, 1):
                lines.append(f"{i}. {name}: £{amt:,.2f}")
            lines.append(f"\nTotal earnings logged: £{earnings_total:,.2f}")
        else:
            lines.append(f"Earnings logged: £{earnings_total:,.2f}")
        lines.append("\nAdd mileage if you want your mileage deduction and estimated "
                     "real take-home.")
        return "\n".join(lines)

    # Mileage-only (§3/§4).
    if mileage and not income:
        lines = [f"{title} mileage record\n", period_line, "",
                 f"Miles logged: {miles_total:,.0f}"]
        used = {vt: m for vt, m in by_vehicle.items() if m > 0}
        if len(used) > 1:
            for vt, m in sorted(used.items(), key=lambda x: -x[1]):
                lines.append(f"  {tax.label(vt)}: {m:,.0f} mi — "
                             f"£{tax.mileage_deduction(m, vt):,.2f} deduction")
        else:
            vt = next(iter(used), tax.normalise_vehicle(user.vehicle_type))
            lines.append(f"Vehicle: {tax.label(vt)}")
        lines.append(f"Mileage deduction captured: £{deduction:,.2f}")
        if user.tax_rate:
            lines.append(f"Estimated tax benefit: up to ~£{benefit:,.2f}")
        if streak > 1:
            lines.append(f"\nStreak: {streak} {streak_unit}s logged 🔥")
        lines.append("\nAdd earnings if you want your estimated real take-home.")
        return "\n".join(lines)

    # Full summary (§1/§2): lead with estimated real take-home.
    estimated_tax = max(0.0, earnings_total - deduction) * (user.tax_rate or 0.0)
    take_home = earnings_total - estimated_tax
    lines = [f"{title} courier summary\n", period_line, ""]

    # Mixed-period warning (§10): mileage and earnings on different frequencies.
    freqs = {r.entry_frequency for r in (income + mileage) if r.entry_frequency}
    if len(freqs) > 1:
        lines.append("⚠️ Your records use different periods (mileage vs earnings), "
                     "so the take-home estimate is less precise.\n")

    lines.append(f"Estimated real take-home: ~£{take_home:,.2f}\n")
    lines.append(f"Earnings logged: £{earnings_total:,.2f}")
    lines.append(f"Miles logged: {miles_total:,.0f}")
    used = {vt: m for vt, m in by_vehicle.items() if m > 0}
    if len(used) > 1:
        for vt, m in sorted(used.items(), key=lambda x: -x[1]):
            lines.append(f"  {tax.label(vt)}: {m:,.0f} mi — "
                         f"£{tax.mileage_deduction(m, vt):,.2f}")
    lines.append(f"Mileage deduction captured: £{deduction:,.2f}")
    if user.tax_rate:
        lines.append(f"Estimated tax benefit: up to ~£{benefit:,.2f}")
    if expenses:
        lines.append("\nExpenses logged for accountant review:")
        for r in expenses:
            lines.append(f"• {r.platform_or_vendor or r.category}: £{(r.amount or 0):,.2f}")
    if streak > 1:
        lines.append(f"\nStreak: {streak} {streak_unit}s logged 🔥")
    lines.append("\nEstimates only, based on what you've confirmed — not tax advice.")
    return "\n".join(lines)


# --- Flow G: Excel record pack (standard export) -----------------------------

def _records_for_export(db: Session, user_id: int, start: str | None, end: str | None):
    rows = _exportable(db, user_id, include_review=True)
    if start and end:
        rows = [r for r in rows if start <= (r.record_date or "") <= end]
    return rows


def build_xlsx(db: Session, user_id: int, user: User | None = None,
               start: str | None = None, end: str | None = None) -> bytes:
    """Standard Flow G export: one workbook with the six required tabs."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    rows = _records_for_export(db, user_id, start, end)
    income = [r for r in rows if r.record_type == "income"]
    mileage = [r for r in rows if r.record_type == "mileage"]
    expenses = [r for r in rows if r.record_type == "expense"
                and r.confirmation_status != "review_required"]
    review = [r for r in rows if r.confirmation_status == "review_required"]

    wb = Workbook()
    bold = Font(bold=True)

    def sheet(name, headers, data_rows):
        ws = wb.create_sheet(name)
        ws.append(headers)
        for c in ws[1]:
            c.font = bold
        for row in data_rows:
            ws.append(row)
        return ws

    wb.remove(wb.active)  # drop default sheet

    period = f"{start or 'all'} to {end or 'all'}"
    rate = user.tax_rate if user else None
    sheet("00_Assumptions", ["item", "value"], [
        ["export_period", period],
        ["method", "simplified mileage"],
        ["tax_estimate_rate", f"{(rate or 0) * 100:.0f}%" if rate is not None else "n/a"],
        ["note", "Record pack for review — not a completed tax return."],
        ["note", "Review-only items are excluded from totals."],
    ])

    sheet("01_Income", ["period_start", "period_end", "entry_frequency", "platform",
                        "amount_gbp", "source_type", "status", "confirmed_at"],
          [[r.period_start or "", r.period_end or "", r.entry_frequency or "",
            r.platform_or_vendor, f"{(r.amount or 0):.2f}", r.source_type,
            r.confirmation_status, str(r.confirmed_at or "")] for r in income])

    sheet("02_Mileage", ["period_start", "period_end", "entry_frequency", "vehicle_type",
                         "miles", "mileage_rate_note", "calculated_deduction",
                         "status", "confirmed_at"],
          [[r.period_start or "", r.period_end or "", r.entry_frequency or "",
            r.vehicle_type or "", f"{(r.miles or 0):.1f}", tax.label(r.vehicle_type),
            f"{tax.mileage_deduction(r.miles or 0, r.vehicle_type):.2f}",
            r.confirmation_status, str(r.confirmed_at or "")] for r in mileage])

    sheet("03_NonVehicleExpenses", ["date", "description", "category_guess",
                                    "amount_gbp", "source_type", "status", "confirmed_at"],
          [[r.record_date, r.platform_or_vendor, r.category, f"{(r.amount or 0):.2f}",
            r.source_type, r.confirmation_status, str(r.confirmed_at or "")]
           for r in expenses])

    sheet("04_ReviewOnly", ["date", "description", "amount_gbp", "reason_for_review",
                            "included_in_total", "status"],
          [[r.record_date, r.platform_or_vendor or r.category, f"{(r.amount or 0):.2f}",
            r.notes, "no", r.confirmation_status] for r in review])

    earnings_total = sum(r.amount or 0 for r in income)
    by_vehicle = _miles_by_vehicle(mileage, user.vehicle_type if user else None)
    miles_total = sum(by_vehicle.values())
    deduction = sum(tax.mileage_deduction(m, vt) for vt, m in by_vehicle.items())
    benefit = (deduction * rate) if rate else 0.0
    expense_total = sum(r.amount or 0 for r in expenses)
    sheet("05_Summary", ["item", "value"], [
        ["export_period", period],
        ["earnings_total", f"{earnings_total:.2f}"],
        ["miles_total", f"{miles_total:.0f}"],
        ["mileage_deduction_total", f"{deduction:.2f}"],
        ["estimated_tax_benefit", f"{benefit:.2f}"],
        ["expense_total_for_review_excluded", f"{expense_total:.2f}"],
        ["review_only_count", str(len(review))],
    ])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def earnings_summary(db: Session, user: User) -> str:
    """Flow D summary after confirming earnings: this-week totals per platform."""
    import datetime as dt
    week_ago = (dt.date.today() - dt.timedelta(days=7)).isoformat()
    week_rows = [r for r in _exportable(db, user.id) if (r.record_date or "") >= week_ago]

    by_platform: dict[str, float] = {}
    for r in week_rows:
        if r.record_type != "income":
            continue
        name = r.platform_or_vendor or "Other"
        by_platform[name] = by_platform.get(name, 0.0) + (r.amount or 0.0)

    # Period line from the most recent income row, if available.
    period = ""
    income_rows = [r for r in week_rows
                   if r.record_type == "income" and r.period_start and r.period_end]
    if income_rows:
        r = income_rows[-1]
        period = f"\nPeriod: {r.period_start} – {r.period_end}"

    lines = [f"Earnings added ✅{period}\n", "This period so far:"]
    for name, amt in sorted(by_platform.items(), key=lambda x: -x[1]):
        lines.append(f"{name}: £{amt:,.2f}")
    total = sum(by_platform.values())
    lines.append(f"\nTotal earnings logged: £{total:,.2f}")

    # Only nudge for mileage if none has been logged this week — otherwise the
    # user just confirmed it and being asked again is confusing.
    has_mileage = any(r.record_type == "mileage" for r in week_rows)
    if not has_mileage:
        lines.append("\nAdd mileage if you want your mileage deduction and estimated "
                     "real take-home.")
    return "\n".join(lines)


def expense_summary(db: Session, user: User) -> str:
    """Flow E1 summary after confirming expenses: this-week expenses listed."""
    import datetime as dt
    week_ago = (dt.date.today() - dt.timedelta(days=7)).isoformat()
    rows = [
        r for r in _exportable(db, user.id)
        if r.record_type == "expense" and (r.record_date or "") >= week_ago
    ]
    lines = ["Expense added ✅\n", "This week's expenses for accountant review:"]
    total = 0.0
    for r in sorted(rows, key=lambda x: x.id):
        label = r.platform_or_vendor or r.category or "Expense"
        total += r.amount or 0.0
        lines.append(f"{label}: £{(r.amount or 0):,.2f}")
    if len(rows) > 1:
        lines.append(f"\nTotal expenses logged: £{total:,.2f}")
    return "\n".join(lines)


def vehicles_overview(db: Session, user: User) -> str:
    """The 'vehicles' command: a per-vehicle tab of miles and deduction."""
    by_vehicle = _miles_by_vehicle(_exportable(db, user.id), user.vehicle_type)
    current = tax.normalise_vehicle(user.vehicle_type)
    by_vehicle.setdefault(current, 0.0)  # always show the active vehicle

    lines = ["Your vehicles:"]
    for vt, miles in sorted(by_vehicle.items(), key=lambda x: -x[1]):
        deduction = tax.mileage_deduction(miles, vt)
        mark = "  ← current" if vt == current else ""
        lines.append(
            f"• {tax.emoji(vt)} {tax.label(vt).capitalize()} — "
            f"{miles:,.0f} miles (£{deduction:,.2f}){mark}"
        )
    lines.append("\nSwitch with \"use car\", \"use motorbike\" or \"use bike\".")
    return "\n".join(lines)
